# Source assistance: OpenAI ChatGPT, 2026-08-22 (AI-05).

from collections.abc import Sequence
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

_HEADER_ROW = 7
_CENTERED_COLUMNS = {1, 2, 3, 8, 11, 12, 13}

_COLUMN_WIDTHS = {
    "A": 8,
    "B": 16,
    "C": 12,
    "D": 32,
    "E": 30,
    "F": 42,
    "G": 38,
    "H": 12,
    "I": 24,
    "J": 28,
    "K": 14,
    "L": 12,
    "M": 26,
}


def build_xlsx(
    module_id: int,
    rows: list[list[str | int]],
    headers: Sequence[str],
) -> bytes:
    """Build the formatted XLSX export."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Approved Test Cases"
    worksheet.sheet_view.showGridLines = False

    cell_border = _build_cell_border()
    _write_title(worksheet)
    _write_summary(worksheet, module_id, len(rows), cell_border)
    _write_header(worksheet, headers, cell_border)
    _write_data(worksheet, rows, cell_border)
    _configure_sheet(worksheet)

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _build_cell_border() -> Border:
    thin_side = Side(style="thin", color="B7C9D6")
    return Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side,
    )


def _write_title(worksheet: Worksheet) -> None:
    worksheet.merge_cells("A1:M1")

    title_cell = worksheet["A1"]
    title_cell.value = "TEST CASE EXPORT"
    title_cell.font = Font(bold=True, size=16, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor="17365D")
    title_cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    worksheet.row_dimensions[1].height = 28


def _write_summary(
    worksheet: Worksheet,
    module_id: int,
    total: int,
    cell_border: Border,
) -> None:
    summary_rows = (
        ("Module ID", module_id),
        ("Status Filter", "APPROVED"),
        ("Total Test Cases", total),
    )

    for row_index, (label, value) in enumerate(summary_rows, start=3):
        label_cell = worksheet.cell(
            row=row_index,
            column=1,
            value=label,
        )
        value_cell = worksheet.cell(
            row=row_index,
            column=2,
            value=value,
        )

        label_cell.font = Font(bold=True)
        label_cell.fill = PatternFill("solid", fgColor="D9EAF7")
        label_cell.border = cell_border
        value_cell.border = cell_border
        label_cell.alignment = Alignment(vertical="center")
        value_cell.alignment = Alignment(vertical="center")


def _write_header(
    worksheet: Worksheet,
    headers: Sequence[str],
    cell_border: Border,
) -> None:
    for column_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(
            row=_HEADER_ROW,
            column=column_index,
            value=header,
        )
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = cell_border

    worksheet.row_dimensions[_HEADER_ROW].height = 32


def _write_data(
    worksheet: Worksheet,
    rows: list[list[str | int]],
    cell_border: Border,
) -> None:
    for row_index, row in enumerate(
        rows,
        start=_HEADER_ROW + 1,
    ):
        _write_data_row(
            worksheet,
            row_index,
            row,
            cell_border,
        )
        _style_status_cell(worksheet, row_index)
        worksheet.row_dimensions[row_index].height = 72


def _write_data_row(
    worksheet: Worksheet,
    row_index: int,
    row: list[str | int],
    cell_border: Border,
) -> None:
    for column_index, value in enumerate(row, start=1):
        cell = worksheet.cell(
            row=row_index,
            column=column_index,
            value=value,
        )
        cell.border = cell_border

        horizontal = "center" if column_index in _CENTERED_COLUMNS else "left"

        cell.alignment = Alignment(
            horizontal=horizontal,
            vertical="top",
            wrap_text=True,
        )


def _style_status_cell(
    worksheet: Worksheet,
    row_index: int,
) -> None:
    status_cell = worksheet.cell(
        row=row_index,
        column=11,
    )
    status_cell.font = Font(
        bold=True,
        color="006100",
    )
    status_cell.fill = PatternFill(
        "solid",
        fgColor="E2F0D9",
    )


def _configure_sheet(worksheet: Worksheet) -> None:
    for column, width in _COLUMN_WIDTHS.items():
        worksheet.column_dimensions[column].width = width

    worksheet.freeze_panes = "A8"
    worksheet.auto_filter.ref = f"A{_HEADER_ROW}:M{worksheet.max_row}"

    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.print_title_rows = f"1:{_HEADER_ROW}"
