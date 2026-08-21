# Source assistance: OpenAI ChatGPT, 2026-08-21 (AI-05).

from datetime import UTC, datetime
from io import BytesIO
from types import SimpleNamespace
from unittest.mock import AsyncMock

import pytest
from app.auth.schemas import CurrentUser
from app.common.constants import AuditAction, ErrorCode, Priority, UserRole
from app.common.constants import TestCaseExportFormat as ExportFormat
from app.common.constants import TestCaseStatus as CaseStatus
from app.common.exceptions import AppError
from app.testcases.export_service import TestCaseExportService as ExportService
from openpyxl import load_workbook


def make_case(*, case_id: int = 10, created_by: int = 7, summary: str = "Add product to cart") -> SimpleNamespace:
    return SimpleNamespace(
        id=case_id,
        requirement_id=2,
        module_id=3,
        summary=summary,
        preconditions=["User is signed in"],
        steps=["Open product", "Add product to cart"],
        expected_result="Product is added",
        priority=Priority.HIGH,
        test_techniques=["equivalence_partitioning"],
        review_note="Reviewed",
        status=CaseStatus.APPROVED,
        created_by=created_by,
        created_at=datetime(2026, 8, 21, 14, 0, tzinfo=UTC),
    )


def build_service(*, test_cases=None, module_exists: bool = True):
    session = SimpleNamespace(commit=AsyncMock())
    repository = SimpleNamespace(list_approved_for_export=AsyncMock(return_value=test_cases or []))
    modules = SimpleNamespace(get_by_id=AsyncMock(return_value=SimpleNamespace(id=3) if module_exists else None))
    audits = SimpleNamespace(create=AsyncMock())
    service = ExportService(session, repository, modules, audits)
    return service, session, repository, modules, audits


@pytest.mark.asyncio
async def test_qa_exports_only_owned_approved_records_and_audits_csv() -> None:
    service, session, repository, _, audits = build_service(test_cases=[make_case()])

    exported = await service.export_approved_test_cases(
        3,
        ExportFormat.CSV,
        CurrentUser(id=7, role=UserRole.QA),
    )

    repository.list_approved_for_export.assert_awaited_once_with(module_id=3, owner_id=7)
    assert exported.filename == "test-cases-module-3.csv"
    assert exported.content.startswith(b"\xef\xbb\xbf")
    assert b"Add product to cart" in exported.content
    audit = audits.create.await_args.args[0]
    assert audit.action == AuditAction.EXPORT_TEST_CASES
    assert audit.entity_id == 3
    assert audit.after_state["test_case_ids"] == [10]
    session.commit.assert_awaited_once()


@pytest.mark.asyncio
async def test_manager_export_is_not_limited_to_record_owner() -> None:
    service, _, repository, _, _ = build_service(test_cases=[make_case(created_by=99)])

    await service.export_approved_test_cases(
        3,
        ExportFormat.CSV,
        CurrentUser(id=2, role=UserRole.MANAGER),
    )

    repository.list_approved_for_export.assert_awaited_once_with(module_id=3, owner_id=None)


@pytest.mark.asyncio
async def test_admin_cannot_export_when_srs_limits_export_role() -> None:
    service, session, repository, modules, audits = build_service(test_cases=[make_case()])

    with pytest.raises(AppError) as exc_info:
        await service.export_approved_test_cases(
            3,
            ExportFormat.CSV,
            CurrentUser(id=1, role=UserRole.ADMIN),
        )

    assert exc_info.value.code == ErrorCode.FORBIDDEN_ROLE
    assert exc_info.value.status_code == 403
    modules.get_by_id.assert_not_awaited()
    repository.list_approved_for_export.assert_not_awaited()
    audits.create.assert_not_awaited()
    session.commit.assert_not_awaited()


@pytest.mark.asyncio
async def test_missing_module_returns_not_found() -> None:
    service, _, repository, _, _ = build_service(module_exists=False)

    with pytest.raises(AppError) as exc_info:
        await service.export_approved_test_cases(
            99,
            ExportFormat.CSV,
            CurrentUser(id=7, role=UserRole.QA),
        )

    assert exc_info.value.code == ErrorCode.MODULE_NOT_FOUND
    assert exc_info.value.status_code == 404
    repository.list_approved_for_export.assert_not_awaited()


@pytest.mark.asyncio
async def test_empty_approved_set_returns_unprocessable_entity() -> None:
    service, session, _, _, audits = build_service(test_cases=[])

    with pytest.raises(AppError) as exc_info:
        await service.export_approved_test_cases(
            3,
            ExportFormat.CSV,
            CurrentUser(id=7, role=UserRole.QA),
        )

    assert exc_info.value.code == ErrorCode.VALIDATION_ERROR
    assert exc_info.value.status_code == 422
    audits.create.assert_not_awaited()
    session.commit.assert_not_awaited()


@pytest.mark.asyncio
async def test_non_approved_record_is_blocked_before_file_creation() -> None:
    case = make_case()
    case.status = CaseStatus.IN_REVIEW
    service, session, _, _, audits = build_service(test_cases=[case])

    with pytest.raises(AppError) as exc_info:
        await service.export_approved_test_cases(
            3,
            ExportFormat.CSV,
            CurrentUser(id=7, role=UserRole.QA),
        )

    assert exc_info.value.code == ErrorCode.CONFLICT
    assert exc_info.value.status_code == 409
    audits.create.assert_not_awaited()
    session.commit.assert_not_awaited()


@pytest.mark.asyncio
async def test_xlsx_export_is_valid_and_formula_text_is_neutralized() -> None:
    service, _, _, _, _ = build_service(test_cases=[make_case(summary='=HYPERLINK("https://example.com")')])

    exported = await service.export_approved_test_cases(
        3,
        ExportFormat.XLSX,
        CurrentUser(id=7, role=UserRole.QA),
    )

    workbook = load_workbook(BytesIO(exported.content), read_only=True, data_only=False)
    worksheet = workbook["Test Cases"]
    assert worksheet.cell(row=2, column=4).value == '\'=HYPERLINK("https://example.com")'
    assert worksheet.cell(row=2, column=11).value == "approved"
