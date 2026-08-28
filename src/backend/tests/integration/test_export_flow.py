# Source assistance: OpenAI ChatGPT, 2026-08-28 (AI-05).

from collections.abc import Callable
from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

DatabaseQuery = Callable[[str, dict | None], list[dict]]
Headers = dict[str, str]


def create_export_data(
    client: TestClient,
    login: Callable[[str], Headers],
) -> tuple[int, int, Headers]:
    manager_headers = login("manager.integration@example.com")

    module_response = client.post(
        "/api/v1/modules",
        headers=manager_headers,
        json={"name": "Export Integration", "parent_id": None},
    )
    assert module_response.status_code == 201
    module_id = module_response.json()["id"]

    qa_headers = login("qa.integration@example.com")
    requirement_response = client.post(
        "/api/v1/requirements",
        headers=qa_headers,
        json={
            "module_id": module_id,
            "content": ("Nguoi dung co the tao booking khi tat ca thong tin bat buoc hop le."),
            "acceptance_criteria": ("He thong tao booking va luu thong tin thanh cong."),
        },
    )
    assert requirement_response.status_code == 201

    return module_id, requirement_response.json()["id"], qa_headers


def seed_export_cases(
    database_write: DatabaseQuery,
    module_id: int,
    requirement_id: int,
) -> tuple[int, int]:
    rows = database_write(
        """
        INSERT INTO test_cases (
            requirement_id, module_id, summary, preconditions,
            steps, expected_result, priority, test_techniques,
            status, created_by, lock_version, tags
        )
        SELECT
            :requirement_id, :module_id, item.summary,
            '["Nguoi dung da dang nhap"]'::json,
            '["Nhap thong tin", "Bam Dat phong"]'::json,
            'Booking duoc tao thanh cong',
            'high'::test_case_priority,
            '["equivalence_partitioning"]'::json,
            item.status::test_case_status,
            u.id, 1, '[]'::json
        FROM users u
        CROSS JOIN (
            VALUES
                ('Approved Booking', 'approved'),
                ('Draft Booking', 'draft')
        ) AS item(summary, status)
        WHERE u.email = 'qa.integration@example.com'
        RETURNING id, summary
        """,
        {
            "requirement_id": requirement_id,
            "module_id": module_id,
        },
    )

    ids = {row["summary"]: row["id"] for row in rows}
    return ids["Approved Booking"], ids["Draft Booking"]


def assert_export_content(
    response,
    export_format: str,
) -> None:
    assert response.status_code == 200
    assert "attachment" in response.headers["content-disposition"]

    if export_format == "csv":
        content = response.content.decode("utf-8-sig")
        assert "Approved Booking" in content
        assert "Draft Booking" not in content
        return

    workbook = load_workbook(BytesIO(response.content))
    values = [str(cell) for row in workbook.active.iter_rows(values_only=True) for cell in row if cell is not None]
    assert "Approved Booking" in values
    assert "Draft Booking" not in values


def assert_export_evidence(
    database_rows: DatabaseQuery,
    approved_id: int,
    draft_id: int,
    module_id: int,
) -> None:
    rows = database_rows(
        """
        SELECT id, status::text AS status
        FROM test_cases
        WHERE id IN (:approved_id, :draft_id)
        """,
        {
            "approved_id": approved_id,
            "draft_id": draft_id,
        },
    )
    statuses = {row["id"]: row["status"] for row in rows}
    assert statuses[approved_id] == "exported"
    assert statuses[draft_id] == "draft"

    audits = database_rows(
        """
        SELECT COUNT(*) AS total
        FROM audit_logs
        WHERE entity_type = 'test_case_export'
          AND entity_id = :module_id
          AND action::text = 'export_test_cases'
        """,
        {"module_id": module_id},
    )
    assert audits[0]["total"] == 1


@pytest.mark.parametrize("export_format", ["csv", "xlsx"])
def test_export_chi_lay_approved(
    export_format: str,
    client: TestClient,
    login: Callable[[str], Headers],
    database_write: DatabaseQuery,
    database_rows: DatabaseQuery,
) -> None:
    module_id, requirement_id, qa_headers = create_export_data(client, login)
    approved_id, draft_id = seed_export_cases(
        database_write,
        module_id,
        requirement_id,
    )

    response = client.get(
        "/api/v1/test-cases/export",
        headers=qa_headers,
        params={
            "module_id": module_id,
            "format": export_format,
        },
    )

    assert_export_content(
        response,
        export_format,
    )
    assert_export_evidence(
        database_rows,
        approved_id,
        draft_id,
        module_id,
    )


def test_export_khong_co_approved_bi_chan(
    client: TestClient,
    login: Callable[[str], Headers],
    database_write: DatabaseQuery,
) -> None:
    module_id, requirement_id, qa_headers = create_export_data(
        client,
        login,
    )

    database_write(
        """
        INSERT INTO test_cases (
            requirement_id, module_id, summary, preconditions,
            steps, expected_result, priority, test_techniques,
            status, created_by, lock_version, tags
        )
        SELECT
            :requirement_id, :module_id, 'Draft only',
            '[]'::json, '["Thuc hien test"]'::json,
            'Khong duoc export',
            'medium'::test_case_priority,
            '[]'::json,
            'draft'::test_case_status,
            u.id, 1, '[]'::json
        FROM users u
        WHERE u.email = 'qa.integration@example.com'
        RETURNING id
        """,
        {
            "requirement_id": requirement_id,
            "module_id": module_id,
        },
    )

    response = client.get(
        "/api/v1/test-cases/export",
        headers=qa_headers,
        params={
            "module_id": module_id,
            "format": "csv",
        },
    )

    assert response.status_code == 422
