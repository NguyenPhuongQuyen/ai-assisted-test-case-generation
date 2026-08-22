# Source assistance: OpenAI ChatGPT, 2026-08-22 (AI-05).

import csv
from dataclasses import dataclass
from io import BytesIO, StringIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy.ext.asyncio import AsyncSession

from app.audit.models import AuditLog
from app.audit.repository import AuditLogRepository
from app.auth.schemas import CurrentUser
from app.common.constants import (
    AuditAction,
    ErrorCode,
    TestCaseExportFormat,
    TestCaseStatus,
    UserRole,
)
from app.common.exceptions import AppError
from app.modules.repository import ModuleRepository
from app.testcases.models import TestCase
from app.testcases.repository import TestCaseRepository


@dataclass(frozen=True, slots=True)
class ExportedTestCaseFile:
    content: bytes
    media_type: str
    filename: str


class TestCaseExportService:
    _EXPORT_ROLES = {UserRole.QA, UserRole.MANAGER}

    # Machine-friendly headers for CSV.
    _HEADERS = (
        "id",
        "requirement_id",
        "module_id",
        "summary",
        "preconditions",
        "steps",
        "expected_result",
        "priority",
        "test_techniques",
        "review_note",
        "status",
        "created_by",
        "created_at",
    )

    # Human-friendly headers for XLSX.
    _DISPLAY_HEADERS = (
        "ID",
        "Requirement ID",
        "Module ID",
        "Summary",
        "Preconditions",
        "Steps",
        "Expected Result",
        "Priority",
        "Test Techniques",
        "Review Note",
        "Status",
        "Created By",
        "Created At",
    )

    def __init__(
        self,
        session: AsyncSession,
        test_cases: TestCaseRepository,
        modules: ModuleRepository,
        audits: AuditLogRepository,
    ) -> None:
        self._session = session
        self._test_cases = test_cases
        self._modules = modules
        self._audits = audits

    async def export_approved_test_cases(
        self,
        module_id: int,
        export_format: TestCaseExportFormat,
        current_user: CurrentUser,
    ) -> ExportedTestCaseFile:
        """Export approved test cases for one module after role and record authorization."""
        self._require_export_role(current_user)

        if await self._modules.get_by_id(module_id) is None:
            raise AppError(
                ErrorCode.MODULE_NOT_FOUND,
                "Không tìm thấy module đã chọn.",
                404,
            )

        # BR-05 / BR-07:
        # QA exports only owned APPROVED records.
        # Manager may export the whole module APPROVED set.
        owner_id = current_user.id if current_user.role == UserRole.QA else None

        test_cases = await self._test_cases.list_approved_for_export(
            module_id=module_id,
            owner_id=owner_id,
        )

        if not test_cases:
            raise AppError(
                ErrorCode.VALIDATION_ERROR,
                "Không có test case APPROVED để export.",
                422,
            )

        self._validate_approved_records(test_cases)

        exported_file = self._build_file(
            module_id,
            export_format,
            test_cases,
        )

        # BR-06 / NC-11: export action is auditable.
        await self._record_export(
            module_id,
            export_format,
            current_user,
            test_cases,
        )
        await self._session.commit()

        return exported_file

    @staticmethod
    def _validate_approved_records(test_cases: list[TestCase]) -> None:
        # BR-01 / BR-05:
        # Never export DRAFT / IN_REVIEW / NEEDS_FIX / EXPORTED / REJECTED.
        if any(test_case.status != TestCaseStatus.APPROVED for test_case in test_cases):
            raise AppError(
                ErrorCode.CONFLICT,
                "Chỉ test case APPROVED mới được export.",
                409,
            )

    def _require_export_role(self, current_user: CurrentUser) -> None:
        if current_user.role not in self._EXPORT_ROLES:
            raise AppError(
                ErrorCode.FORBIDDEN_ROLE,
                "Vai trò hiện tại không có quyền export test case.",
                403,
            )

    def _build_file(
        self,
        module_id: int,
        export_format: TestCaseExportFormat,
        test_cases: list[TestCase],
    ) -> ExportedTestCaseFile:
        rows = [self._to_row(test_case) for test_case in test_cases]

        if export_format == TestCaseExportFormat.CSV:
            return ExportedTestCaseFile(
                content=self._build_csv(rows),
                media_type="text/csv; charset=utf-8",
                filename=f"test-cases-module-{module_id}.csv",
            )

        return ExportedTestCaseFile(
            content=self._build_xlsx(module_id, rows),
            media_type=("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
            filename=f"test-cases-module-{module_id}.xlsx",
        )

    async def _record_export(
        self,
        module_id: int,
        export_format: TestCaseExportFormat,
        current_user: CurrentUser,
        test_cases: list[TestCase],
    ) -> None:
        await self._audits.create(
            AuditLog(
                user_id=current_user.id,
                action=AuditAction.EXPORT_TEST_CASES,
                entity_type="test_case_export",
                entity_id=module_id,
                before_state=None,
                after_state={
                    "module_id": module_id,
                    "format": export_format.value,
                    "count": len(test_cases),
                    "test_case_ids": [test_case.id for test_case in test_cases],
                },
            )
        )

    @classmethod
    def _build_csv(cls, rows: list[list[str | int]]) -> bytes:
        """
        Build an Excel-friendly UTF-8 CSV.

        Semicolon is used because Vietnamese/European Excel installations
        commonly use comma as decimal separator and semicolon as list separator.
        Quoting is still handled by Python's csv module.
        """
        stream = StringIO(newline="")

        writer = csv.writer(
            stream,
            delimiter=";",
            quotechar='"',
            quoting=csv.QUOTE_MINIMAL,
            lineterminator="\r\n",
        )

        writer.writerow(cls._DISPLAY_HEADERS)

        csv_rows = []
        for row in rows:
            csv_row = list(row)

            # CSV is primarily an interchange format, so keep each
            # Test Case on one physical line for easier Excel/import viewing.
            csv_row[4] = str(csv_row[4]).replace("\r\n", " | ").replace("\n", " | ")
            csv_row[5] = str(csv_row[5]).replace("\r\n", " | ").replace("\n", " | ")

            # Human-friendly enum display without changing stored values.
            csv_row[7] = str(csv_row[7]).upper()
            csv_row[10] = str(csv_row[10]).upper()

            csv_rows.append(csv_row)

        writer.writerows(csv_rows)

        # UTF-8 BOM lets Excel detect Vietnamese Unicode correctly.
        return stream.getvalue().encode("utf-8-sig")

    @classmethod
    def _build_xlsx(
        cls,
        module_id: int,
        rows: list[list[str | int]],
    ) -> bytes:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Approved Test Cases"
        worksheet.sheet_view.showGridLines = False

        # ---------- Palette ----------
        dark_blue = "17365D"
        header_blue = "1F4E78"
        info_blue = "D9EAF7"
        approved_green = "E2F0D9"
        white = "FFFFFF"
        border_color = "B7C9D6"

        thin_side = Side(
            style="thin",
            color=border_color,
        )
        cell_border = Border(
            left=thin_side,
            right=thin_side,
            top=thin_side,
            bottom=thin_side,
        )

        # ---------- Title ----------
        worksheet.merge_cells("A1:M1")
        title_cell = worksheet["A1"]
        title_cell.value = "TEST CASE EXPORT"
        title_cell.font = Font(
            bold=True,
            size=16,
            color=white,
        )
        title_cell.fill = PatternFill(
            "solid",
            fgColor=dark_blue,
        )
        title_cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )
        worksheet.row_dimensions[1].height = 28

        # ---------- Export summary ----------
        summary_rows = (
            ("Module ID", module_id),
            ("Status Filter", "APPROVED"),
            ("Total Test Cases", len(rows)),
        )

        for row_index, (label, value) in enumerate(
            summary_rows,
            start=3,
        ):
            label_cell = worksheet.cell(
                row=row_index,
                column=1,
                value=label,
            )
            value_cell = worksheet.cell(
                row=row_index,
                column=2,
                value=value,
            )

            label_cell.font = Font(bold=True)
            label_cell.fill = PatternFill(
                "solid",
                fgColor=info_blue,
            )

            label_cell.border = cell_border
            value_cell.border = cell_border

            label_cell.alignment = Alignment(
                vertical="center",
            )
            value_cell.alignment = Alignment(
                vertical="center",
            )

        # ---------- Table header ----------
        header_row = 7

        for column_index, header in enumerate(
            cls._DISPLAY_HEADERS,
            start=1,
        ):
            cell = worksheet.cell(
                row=header_row,
                column=column_index,
                value=header,
            )
            cell.font = Font(
                bold=True,
                color=white,
            )
            cell.fill = PatternFill(
                "solid",
                fgColor=header_blue,
            )
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
            cell.border = cell_border

        worksheet.row_dimensions[header_row].height = 32

        # ---------- Data ----------
        for row_index, row in enumerate(
            rows,
            start=header_row + 1,
        ):
            for column_index, value in enumerate(
                row,
                start=1,
            ):
                cell = worksheet.cell(
                    row=row_index,
                    column=column_index,
                    value=value,
                )

                cell.border = cell_border

                if column_index in {
                    1,
                    2,
                    3,
                    8,
                    11,
                    12,
                    13,
                }:
                    cell.alignment = Alignment(
                        horizontal="center",
                        vertical="top",
                        wrap_text=True,
                    )
                else:
                    cell.alignment = Alignment(
                        horizontal="left",
                        vertical="top",
                        wrap_text=True,
                    )

            # APPROVED status cell.
            status_cell = worksheet.cell(
                row=row_index,
                column=11,
            )
            status_cell.font = Font(
                bold=True,
                color="006100",
            )
            status_cell.fill = PatternFill(
                "solid",
                fgColor=approved_green,
            )

            # Give long text enough vertical room without editing manually.
            worksheet.row_dimensions[row_index].height = 72

        # ---------- Column widths ----------
        column_widths = {
            "A": 8,
            "B": 16,
            "C": 12,
            "D": 32,
            "E": 30,
            "F": 42,
            "G": 38,
            "H": 12,
            "I": 24,
            "J": 28,
            "K": 14,
            "L": 12,
            "M": 26,
        }

        for column, width in column_widths.items():
            worksheet.column_dimensions[column].width = width

        # Keep header visible while scrolling.
        worksheet.freeze_panes = "A8"

        # Excel filter on the actual Test Case table.
        worksheet.auto_filter.ref = f"A{header_row}:M{worksheet.max_row}"

        # Friendly printing if the lecturer prints / exports to PDF.
        worksheet.page_setup.orientation = "landscape"
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0
        worksheet.sheet_properties.pageSetUpPr.fitToPage = True
        worksheet.print_title_rows = f"1:{header_row}"

        stream = BytesIO()
        workbook.save(stream)
        return stream.getvalue()

    @classmethod
    def _to_row(
        cls,
        test_case: TestCase,
    ) -> list[str | int]:
        return [
            test_case.id,
            test_case.requirement_id,
            test_case.module_id,
            cls._safe_cell(test_case.summary),
            cls._safe_cell(cls._join_numbered_steps(test_case.preconditions)),
            cls._safe_cell(cls._join_numbered_steps(test_case.steps)),
            cls._safe_cell(test_case.expected_result),
            test_case.priority.value,
            cls._safe_cell(", ".join(test_case.test_techniques)),
            cls._safe_cell(test_case.review_note or ""),
            test_case.status.value,
            test_case.created_by,
            test_case.created_at.isoformat(),
        ]

    @staticmethod
    def _join_lines(values: list[str]) -> str:
        return "\n".join(values)

    @staticmethod
    def _join_numbered_steps(values: list[str]) -> str:
        return "\n".join(
            f"{index}. {value}"
            for index, value in enumerate(
                values,
                start=1,
            )
        )

    @staticmethod
    def _safe_cell(value: str) -> str:
        # Security: prevent spreadsheet formula execution.
        if value.startswith(("=", "+", "-", "@")):
            return "'" + value

        return value
